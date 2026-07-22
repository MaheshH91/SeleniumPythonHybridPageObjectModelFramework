from typing import Any, List
import openpyxl


def get_row_count(path: str, sheet_name: str) -> int:
    """Returns total number of rows in the specified worksheet."""
    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    workbook.close()
    return row_count


def get_column_count(path: str, sheet_name: str) -> int:
    """Returns total number of columns in the specified worksheet."""
    workbook = openpyxl.load_workbook(path, read_only=True)
    sheet = workbook[sheet_name]
    col_count = sheet.max_column
    workbook.close()
    return col_count


def get_cell_data(path: str, sheet_name: str, row_number: int, column_number: int) -> Any:
    """Retrieves value of a specific cell."""
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    cell_value = sheet.cell(row=row_number, column=column_number).value
    workbook.close()
    return cell_value


def set_cell_data(path: str, sheet_name: str, row_number: int, column_number: int, data: Any) -> None:
    """Writes data to a specific cell and saves the workbook."""
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)
    workbook.close()


def get_data_from_excel(path: str, sheet_name: str) -> List[List[Any]]:
    """
    Reads all data rows (excluding the header row) and returns a list of lists
    formatted for @pytest.mark.parametrize.
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]

    final_list = []
    # sheet.iter_rows speeds up row extraction compared to nested cell() lookups
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Ignore completely empty rows in Excel
        if any(cell is not None for cell in row):
            final_list.append(list(row))

    workbook.close()
    return final_list